import re
import sys
from PIL import Image
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QTabWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QRadioButton,
    QGroupBox,
    QTextEdit,
    QComboBox,
    QFileDialog,
    QMessageBox
)
from openpyxl import load_workbook
import os
from shutil import copy2
import pandas as pd


class ImageProcessingTool(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("图片处理工具")
        self.resize(800, 600)

        # 创建中心窗口和主布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 先定义 edit_result（统一使用）
        self.edit_result = QTextEdit()

        # 处理结果文本框（统一使用）
        result_group = QGroupBox("处理结果")
        result_layout = QVBoxLayout(result_group)
        
        result_layout.addWidget(self.edit_result)

        # 创建选项卡
        self.tab_widget = QTabWidget()
        self.tab_extract = self.create_extract_tab()
        self.tab_rename = self.create_rename_tab()
        self.tab_compress = self.create_compress_tab()
        self.tab_widget.addTab(self.tab_extract, "提取照片")
        self.tab_widget.addTab(self.tab_rename, "重命名照片")
        self.tab_widget.addTab(self.tab_compress, "压缩照片")

        main_layout.addWidget(self.tab_widget)
        main_layout.addWidget(result_group)  # 👈 将处理结果显示区域加入主界面

    def select_folder(self, line_edit):
        """通用文件夹选择对话框"""
        folder_path = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if folder_path:
            line_edit.setText(folder_path)

    # ----------------------------
    # 提取照片 Tab
    # ----------------------------
    def create_extract_tab(self):
        """创建“提取照片”选项卡"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)

        # 第一步：选择Excel文件
        group_box_excel = QGroupBox("第一步：选择Excel文件")
        layout_excel = QVBoxLayout(group_box_excel)
        self.edit_excel_path = QLineEdit("未选择Excel文件")
        self.btn_select_excel = QPushButton("选择Excel文件")
        layout_excel.addWidget(self.edit_excel_path)
        layout_excel.addWidget(self.btn_select_excel)
        self.label_sheet = QLabel("选择工作表：")
        self.combo_sheet = QComboBox()  # 替换为 QComboBox
        layout_excel.addWidget(self.label_sheet)
        layout_excel.addWidget(self.combo_sheet)
        layout.addWidget(group_box_excel)

        # 绑定事件
        self.btn_select_excel.clicked.connect(self.select_excel_file)

        # 第二步：选择照片文件夹
        group_box_photo_src = QGroupBox("第二步：选择照片文件夹")
        layout_photo_src = QVBoxLayout(group_box_photo_src)
        self.edit_photo_src_path = QLineEdit("未选择照片源文件夹")
        self.btn_select_photo_src = QPushButton("选择源文件夹")
        layout_photo_src.addWidget(self.edit_photo_src_path)
        layout_photo_src.addWidget(self.btn_select_photo_src)
        layout.addWidget(group_box_photo_src)

        # 绑定事件
        self.btn_select_photo_src.clicked.connect(lambda: self.select_folder(self.edit_photo_src_path))

        # 第三步：选择目标文件夹
        group_box_photo_dest = QGroupBox("第三步：选择目标文件夹")
        layout_photo_dest = QVBoxLayout(group_box_photo_dest)
        self.edit_photo_dest_path = QLineEdit("未选择目标文件夹")
        self.btn_select_photo_dest = QPushButton("选择目标文件夹")
        layout_photo_dest.addWidget(self.edit_photo_dest_path)
        layout_photo_dest.addWidget(self.btn_select_photo_dest)
        layout.addWidget(group_box_photo_dest)

        # 绑定事件
        self.btn_select_photo_dest.clicked.connect(lambda: self.select_folder(self.edit_photo_dest_path))

        # 开始提取按钮
        self.btn_start_extract = QPushButton("开始提取照片")
        

        # 添加“全选”项
        self.combo_sheet.addItem("全选")

        # 绑定按钮事件
        self.btn_start_extract.clicked.connect(self.start_extract)

        return tab_widget

    def select_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.edit_excel_path.setText(file_path)
            self.load_excel_sheets(file_path)

    def load_excel_sheets(self, file_path):
        try:
            workbook = load_workbook(filename=file_path)
            sheet_names = workbook.sheetnames
            self.combo_sheet.clear()
            self.combo_sheet.addItems(sheet_names)
            self.combo_sheet.addItem("全选")
        except Exception as e:
            self.combo_sheet.clear()
            print(f"无法读取Excel文件: {e}")

    def start_extract(self):
        """开始提取照片"""
        excel_path = self.edit_excel_path.text()
        sheet_name = self.combo_sheet.currentText()
        photo_src_folder = self.edit_photo_src_path.text()
        photo_dest_folder = self.edit_photo_dest_path.text()

        if not os.path.isfile(excel_path):
            self.edit_result.append("❌ Excel文件路径无效")
            return
        if not os.path.isdir(photo_src_folder):
            self.edit_result.append("❌ 源文件夹路径无效")
            return
        if not os.path.isdir(photo_dest_folder):
            self.edit_result.append("❌ 目标文件夹路径无效")
            return

        try:
            # 读取Excel数据
            if sheet_name == "全选":
                sheets = pd.read_excel(excel_path, sheet_name=None)
                all_ids = set()
                for df in sheets.values():
                    if '身份证号' in df.columns:
                        all_ids.update(df['身份证号'].astype(str).str.strip())
            else:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                if '身份证号' not in df.columns:
                    self.edit_result.append(f"❌ 工作表 {sheet_name} 中没有‘身份证号’列")
                    return
                all_ids = df['身份证号'].astype(str).str.strip().unique()

            success_count = 0
            failed_ids = []

            for idx, id_number in enumerate(all_ids, start=1):
                self.edit_result.append(f"{idx}. 正在处理身份证号：{id_number}")
                matching_files = [f for f in os.listdir(photo_src_folder) if id_number in f]
                image_files = [f for f in matching_files if f.lower().endswith(('.png', '.jpg', '.jpeg'))]

                if len(image_files) == 1:
                    src_file = os.path.join(photo_src_folder, image_files[0])
                    dest_file = os.path.join(photo_dest_folder, image_files[0])
                    copy2(src_file, dest_file)
                    self.edit_result.append(f"✅ 成功复制：{image_files[0]}")
                    success_count += 1
                else:
                    failed_ids.append(id_number)
                    self.edit_result.append(f"❌ 身份证号 {id_number} 匹配到 {len(image_files)} 个文件，请检查命名")

            if failed_ids:
                QMessageBox.warning(self, "部分失败", f"{len(failed_ids)} 个身份证号未能正确匹配照片")
            else:
                QMessageBox.information(self, "完成", f"共提取 {success_count} 张照片，全部成功！")
        except Exception as e:
            self.edit_result.append(f"❌ 出现错误：{str(e)}")
            QMessageBox.critical(self, "错误", str(e))

    # ----------------------------
    # 重命名照片 Tab
    # ----------------------------
    def create_rename_tab(self):
        """创建“重命名照片”选项卡"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)

        # 选择源文件夹
        group_box_src = QGroupBox("选择源文件夹")
        layout_src = QVBoxLayout(group_box_src)
        self.edit_rename_src_path = QLineEdit("未选择源文件夹")
        self.btn_select_rename_src = QPushButton("选择源文件夹")
        layout_src.addWidget(self.edit_rename_src_path)
        layout_src.addWidget(self.btn_select_rename_src)
        layout.addWidget(group_box_src)

        # 绑定事件
        self.btn_select_rename_src.clicked.connect(lambda: self.select_folder(self.edit_rename_src_path))

        # 选择目标文件夹
        group_box_dest = QGroupBox("选择目标文件夹")
        layout_dest = QVBoxLayout(group_box_dest)
        self.edit_rename_dest_path = QLineEdit("未选择目标文件夹")
        self.btn_select_rename_dest = QPushButton("选择目标文件夹")
        layout_dest.addWidget(self.edit_rename_dest_path)
        layout_dest.addWidget(self.btn_select_rename_dest)
        layout.addWidget(group_box_dest)

        # 绑定事件
        self.btn_select_rename_dest.clicked.connect(lambda: self.select_folder(self.edit_rename_dest_path))

        # 开始重命名按钮
        self.btn_start_rename = QPushButton("开始重命名照片")
        layout.addWidget(self.btn_start_rename)
        # 绑定事件
        self.btn_start_rename.clicked.connect(self.start_rename)

        return tab_widget

    def start_rename(self):
        """开始重命名照片"""
        src_folder = self.edit_rename_src_path.text()
        dest_folder = self.edit_rename_dest_path.text()

        if not os.path.isdir(src_folder):
            self.edit_result.append("❌ 源文件夹路径无效")
            return
        if not os.path.isdir(dest_folder):
            self.edit_result.append("❌ 目标文件夹路径无效")
            return

        # 支持的图片格式
        valid_extensions = ('.jpg', '.jpeg', '.png')
        # 正则表达式匹配18位身份证号码
        id_pattern = re.compile(r'.*?(\d{17}[\dXx]).*', re.IGNORECASE)

        success_count = 0
        failed_files = []

        for filename in os.listdir(src_folder):
            file_lower = filename.lower()
            if not file_lower.endswith(valid_extensions):
                continue

            match = id_pattern.match(filename)
            if match:
                id_number = match.group(1).upper()
                ext = os.path.splitext(filename)[1]
                new_name = f"{id_number}{ext}"
                src_path = os.path.join(src_folder, filename)
                dest_path = os.path.join(dest_folder, new_name)

                try:
                    copy2(src_path, dest_path)
                    self.edit_result.append(f"✅ 成功重命名：{filename} → {new_name}")
                    success_count += 1
                except Exception as e:
                    self.edit_result.append(f"❌ 无法复制文件 {filename}: {str(e)}")
                    failed_files.append(filename)
            else:
                self.edit_result.append(f"❌ 未找到身份证号：{filename}")
                failed_files.append(filename)

        if failed_files:
            QMessageBox.warning(self, "部分失败", f"{len(failed_files)} 个文件未能正确重命名，请检查文件名是否包含身份证号")
        else:
            QMessageBox.information(self, "完成", f"共重命名 {success_count} 个文件，全部成功！")

    # ----------------------------
    # 压缩照片 Tab
    # ----------------------------
    def create_compress_tab(self):
        """创建“压缩照片”选项卡（仅支持文件夹）"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)

        # 选择压缩源（仅保留文件夹）
        group_box_compress_src = QGroupBox("压缩源文件夹")
        layout_compress_src = QVBoxLayout(group_box_compress_src)
        self.edit_compress_folder_path = QLineEdit("未选择文件夹")
        self.btn_select_compress_folder = QPushButton("选择文件夹")
        layout_compress_src.addWidget(self.edit_compress_folder_path)
        layout_compress_src.addWidget(self.btn_select_compress_folder)
        layout.addWidget(group_box_compress_src)

        # 绑定事件
        self.btn_select_compress_folder.clicked.connect(lambda: self.select_folder(self.edit_compress_folder_path))

        # 目标文件夹
        group_box_dest = QGroupBox("目标文件夹")
        layout_dest = QVBoxLayout(group_box_dest)
        self.edit_compress_dest_path = QLineEdit("未选择目标文件夹")
        self.btn_select_compress_dest = QPushButton("选择目标文件夹")
        layout_dest.addWidget(self.edit_compress_dest_path)
        layout_dest.addWidget(self.btn_select_compress_dest)
        layout.addWidget(group_box_dest)

        # 绑定事件
        self.btn_select_compress_dest.clicked.connect(lambda: self.select_folder(self.edit_compress_dest_path))

        # 压缩设置
        group_box_settings = QGroupBox("压缩设置")
        layout_settings = QVBoxLayout(group_box_settings)
        self.label_target_size = QLabel("目标大小（KB）：")
        self.edit_target_size = QLineEdit("100")
        layout_settings.addWidget(self.label_target_size)
        layout_settings.addWidget(self.edit_target_size)
        layout.addWidget(group_box_settings)

        # 开始压缩按钮
        self.btn_start_compress = QPushButton("开始压缩照片")  # 👈 先定义按钮
        layout.addWidget(self.btn_start_compress)  # 👈 添加到主布局
        # 绑定事件
        self.btn_start_compress.clicked.connect(self.start_compress)

        return tab_widget
    def start_compress(self):
        """开始压缩照片"""
        src_folder = self.edit_compress_folder_path.text()
        dest_folder = self.edit_compress_dest_path.text()
        target_size_kb_text = self.edit_target_size.text()

        # 检查输入合法性
        if not os.path.isdir(src_folder):
            self.edit_result.append("❌ 源文件夹路径无效")
            return
        if not os.path.isdir(dest_folder):
            self.edit_result.append("❌ 目标文件夹路径无效")
            return
        try:
            target_size_kb = int(target_size_kb_text)
            if target_size_kb <= 0:
                raise ValueError
        except:
            self.edit_result.append("❌ 目标大小必须是正整数")
            return

        success_count = 0
        failed_files = []

        for filename in os.listdir(src_folder):
            file_path = os.path.join(src_folder, filename)
            if not filename.lower().endswith(".jpg"):
                continue

            file_size_kb = os.path.getsize(file_path) // 1024
            if file_size_kb <= target_size_kb:
                self.edit_result.append(f"⏩ 跳过（未超过目标大小）: {filename}")
                continue

            try:
                img = Image.open(file_path)
                original_width, original_height = img.size
                self.edit_result.append(f"📸 处理图片: {filename} | 原始尺寸: {original_width}x{original_height} | 大小: {file_size_kb} KB")

                dest_path = os.path.join(dest_folder, filename)

                quality = 95
                while quality >= 30:
                    img.save(dest_path, "JPEG", quality=quality, optimize=True, progressive=True)
                    compressed_size_kb = os.path.getsize(dest_path) // 1024
                    self.edit_result.append(f"⚙️ 尝试质量={quality} → 大小: {compressed_size_kb} KB")
                    if compressed_size_kb <= target_size_kb:
                        self.edit_result.append(f"✅ 成功压缩: {filename} → {compressed_size_kb} KB (质量: {quality})")
                        success_count += 1
                        break
                    quality -= 10

                if quality < 30:
                    # 尝试缩放后再压缩
                    min_side = min(original_width, original_height)
                    self.edit_result.append(f"📐 图像质量已降至最低({quality}), 准备尝试缩放尺寸...")

                    if min_side > 1080:
                        scale_ratio = 1080 / min_side
                        new_size = (int(original_width * scale_ratio), int(original_height * scale_ratio))
                        self.edit_result.append(f"📐 缩放比例: {scale_ratio:.2f} → 新尺寸: {new_size}")

                        img = img.resize(new_size, Image.LANCZOS)
                        quality = 95
                        while quality >= 30:
                            img.save(dest_path, "JPEG", quality=quality, optimize=True, progressive=True)
                            compressed_size_kb = os.path.getsize(dest_path) // 1024
                            self.edit_result.append(f"⚙️ 缩放后尝试质量={quality} → 大小: {compressed_size_kb} KB")
                            if compressed_size_kb <= target_size_kb:
                                self.edit_result.append(f"✅ 成功压缩（缩放后）: {filename} → {compressed_size_kb} KB (质量: {quality})")
                                success_count += 1
                                break
                            quality -= 10

                        if quality < 30:
                            raise Exception("无法压缩至目标大小（缩放后仍失败）")
                    else:
                        raise Exception("图像尺寸太小，无法通过缩放进一步压缩")

            except Exception as e:
                self.edit_result.append(f"❌ 压缩失败: {filename} → {str(e)}")
                failed_files.append(filename)

        if failed_files:
            QMessageBox.warning(self, "部分失败", f"{len(failed_files)} 张图片未能成功压缩")
        else:
            QMessageBox.information(self, "完成", f"共压缩 {success_count} 张图片，全部成功！")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ImageProcessingTool()
    window.show()
    sys.exit(app.exec_())